"""Scheduled task data model, schedule parsing, and Excel I/O."""

from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Optional

import pandas as pd
from openpyxl import Workbook


@dataclass
class ScheduledTask:
    """A single scheduled task loaded from the tasks Excel file."""

    name: str
    schedule: str
    description: str
    last_run: Optional[datetime] = None
    added_by: str = "user"

    # Computed from schedule text via parse_schedule():
    _interval_seconds: Optional[int] = field(default=None, repr=False)
    _daily_time_utc: Optional[tuple[int, int]] = field(default=None, repr=False)

    def is_due(self, now: datetime | None = None) -> bool:
        """Check whether this task should run at the given time.

        Args:
            now: Reference time for the check. Defaults to the current
                UTC time if not provided.

        Returns:
            True if the task is due for execution, False otherwise.
        """
        now = now or datetime.now(timezone.utc)

        if self._interval_seconds is not None:
            if self.last_run is None:
                return True
            elapsed = (now - self.last_run).total_seconds()
            return elapsed >= self._interval_seconds

        if self._daily_time_utc is not None:
            target_h, target_m = self._daily_time_utc
            if now.hour == target_h and now.minute == target_m:
                if self.last_run is None:
                    return True
                return self.last_run.date() < now.date()

        return False


# ── Schedule Parsing ─────────────────────────────────────────

# Timezone abbreviation -> UTC offset (hours).  Uses standard time;
# DST adds ~1 hour but we accept the approximation.
_TZ_OFFSETS = {
    "et": -5, "est": -5, "edt": -4,
    "ct": -6, "cst": -6, "cdt": -5,
    "pt": -8, "pst": -8, "pdt": -7,
    "mt": -7, "mst": -7, "mdt": -6,
    "utc": 0, "gmt": 0,
}


def parse_schedule(text: str) -> tuple[Optional[int], Optional[tuple[int, int]]]:
    """Parse free-text schedule into an interval or daily-time pair.

    Exactly one of the two return values will be non-None.

    Supported patterns::

        "Once per hour" / "Every hour" / "Hourly"
        "Every 30 minutes" / "Once per 15 minutes"
        "Every 2 hours"
        "1230pm ET daily" / "12:30 PM ET daily"
        "9am daily" / "Daily at 5pm"

    Args:
        text: Human-readable schedule string to parse.

    Returns:
        A tuple of ``(interval_seconds, daily_time_utc)``. For interval
        schedules the first element is the repeat period in seconds and
        the second is None. For daily schedules the first element is None
        and the second is an ``(hour, minute)`` tuple in UTC.

    Raises:
        ValueError: If the schedule text cannot be parsed.
    """
    t = text.strip().lower().rstrip(".")

    # ── Interval patterns ────────────────────────────────
    if re.search(r"\bhourly\b|once per hour\b|every\s+hour\b", t):
        return (3600, None)

    m = re.search(r"(?:every|once per)\s+(\d+)\s*min(?:ute)?s?", t)
    if m:
        return (int(m.group(1)) * 60, None)

    m = re.search(r"(?:every|once per)\s+(\d+)\s*hours?", t)
    if m:
        return (int(m.group(1)) * 3600, None)

    # ── Daily-at-time patterns ───────────────────────────
    # Try "1230pm ET daily" or "12:30 PM ET daily" first
    m = re.search(
        r"(\d{1,4})\s*:?\s*(\d{2})?\s*(am|pm)?\s*"
        r"(?:(et|est|edt|ct|cst|cdt|pt|pst|pdt|mt|mst|mdt|utc|gmt)\s+)?"
        r"(?:daily|every\s*day)",
        t,
    )
    if not m:
        # Try reversed: "daily at 5pm ET"
        m = re.search(
            r"(?:daily|every\s*day)\s+(?:at\s+)?"
            r"(\d{1,4})\s*:?\s*(\d{2})?\s*(am|pm)?\s*"
            r"(et|est|edt|ct|cst|cdt|pt|pst|pdt|mt|mst|mdt|utc|gmt)?",
            t,
        )

    if m:
        raw_hour = m.group(1)
        raw_min = m.group(2)
        ampm = m.group(3)
        tz_abbr = (m.group(4) or "et").lower()

        # Handle compact time like "1230" (no colon)
        if len(raw_hour) > 2 and raw_min is None:
            minute = int(raw_hour[-2:])
            hour = int(raw_hour[:-2])
        else:
            hour = int(raw_hour)
            minute = int(raw_min or 0)

        if ampm == "pm" and hour != 12:
            hour += 12
        elif ampm == "am" and hour == 12:
            hour = 0

        offset = _TZ_OFFSETS.get(tz_abbr, -5)
        utc_hour = (hour - offset) % 24

        return (None, (utc_hour, minute))

    raise ValueError(f"Cannot parse schedule: '{text}'")


# ── Excel I/O ────────────────────────────────────────────────


def load_tasks_from_excel(file_path: str) -> list[ScheduledTask]:
    """Load scheduled tasks from an Excel file.

    Expected columns: TASK NAME, SCHEDULE, DESCRIPTION, LAST RUN, ADDED BY.

    Args:
        file_path: Path to the ``.xlsx`` file. If the file does not exist,
            an empty list is returned.

    Returns:
        List of parsed ``ScheduledTask`` instances. Tasks with unparseable
        schedules are still included (a warning is printed to stdout).
    """
    if not os.path.exists(file_path):
        return []

    df = pd.read_excel(file_path)
    if df.empty:
        return []

    tasks: list[ScheduledTask] = []
    for _, row in df.iterrows():
        name = str(row.get("TASK NAME", "")).strip()
        schedule = str(row.get("SCHEDULE", "")).strip()
        description = str(row.get("DESCRIPTION", "")).strip()
        last_run_raw = row.get("LAST RUN")
        added_by = str(row.get("ADDED BY", "user")).strip()

        if not name:
            continue

        # Parse last_run
        last_run: datetime | None = None
        if pd.notna(last_run_raw) and str(last_run_raw).strip().lower() not in (
            "never", "none", "",
        ):
            try:
                last_run = pd.to_datetime(last_run_raw).to_pydatetime()
                if last_run.tzinfo is None:
                    last_run = last_run.replace(tzinfo=timezone.utc)
            except Exception:
                pass

        task = ScheduledTask(
            name=name,
            schedule=schedule,
            description=description,
            last_run=last_run,
            added_by=added_by or "user",
        )

        try:
            interval, daily_time = parse_schedule(schedule)
            task._interval_seconds = interval
            task._daily_time_utc = daily_time
        except ValueError:
            print(f"Warning: could not parse schedule for task '{name}': {schedule}")

        tasks.append(task)

    return tasks


def save_tasks_to_excel(tasks: list[ScheduledTask], file_path: str) -> None:
    """Save scheduled tasks to an Excel file, creating or overwriting it.

    Args:
        tasks: The task instances to persist.
        file_path: Destination ``.xlsx`` path. The file is created if it
            does not exist, or overwritten if it does.
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None

    headers = ["TASK NAME", "SCHEDULE", "DESCRIPTION", "LAST RUN", "ADDED BY"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, task in enumerate(tasks, start=2):
        ws.cell(row=row_idx, column=1, value=task.name)
        ws.cell(row=row_idx, column=2, value=task.schedule)
        ws.cell(row=row_idx, column=3, value=task.description)
        ws.cell(row=row_idx, column=4,
                value=task.last_run.isoformat() if task.last_run else "never")
        ws.cell(row=row_idx, column=5, value=task.added_by)

    wb.save(file_path)
