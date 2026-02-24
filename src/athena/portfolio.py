# pyright: reportUnknownMemberType=false, reportUnknownArgumentType=false

from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo

from typing import Any, Union

from decimal import Decimal

from collections import defaultdict

from enum import Enum

import warnings

from .currency import Currency, ExchangeRateManager, FederalReserveExchangeRateManager
from .pricingdata import PricingDataManager, YFinancePricingDataManager

import json

# Default timezone for transactions without timezone info
NYC_TIMEZONE = ZoneInfo("America/New_York")


def _normalize_transaction_datetime(dt: datetime) -> tuple[datetime, bool, bool]:
    """
    Normalize a transaction datetime to ensure it has timezone information.

    If timezone is missing, assumes NYC timezone.
    If time is midnight (00:00:00), assumes 12:00 PM NYC time as time may be missing.

    Args:
        dt: The datetime to normalize.

    Returns:
        A tuple of (normalized_datetime, time_was_missing, timezone_was_missing).
    """
    time_was_missing = False
    timezone_was_missing = False

    # Check if time appears to be missing (midnight with no microseconds)
    if dt.hour == 0 and dt.minute == 0 and dt.second == 0 and dt.microsecond == 0:
        dt = dt.replace(hour=12, minute=0, second=0, microsecond=0)
        time_was_missing = True

    # Check if timezone is missing
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=NYC_TIMEZONE)
        timezone_was_missing = True

    return dt, time_was_missing, timezone_was_missing

import pandas as pd
from openpyxl import Workbook

class TransactionType(Enum):
    """Enumeration of supported portfolio transaction types."""

    BUY = "BUY"
    SELL = "SELL"
    CASH_IN = "CASH_IN"
    CASH_OUT = "CASH_OUT"
    DIVIDEND = "DIVIDEND"
    INTEREST = "INTEREST"
    FEE = "FEE"
    CURRENCY_EXCHANGE = "CURRENCY_EXCHANGE" # How this works: the SYMBOL is the target currency (e.g. "HKD") and the CURRENCY is the source currency (e.g., "USD"). The price is how much of the CURRENCY it costs to buy 1 unit of the target (SYMBOL) currency.

class Position():
    """A snapshot of a held asset at a specific point in time."""

    def __init__(self, symbol:str, quantity:Decimal, position_datetime:datetime, total_value:Decimal, book_value:Decimal | None = None, unit_price:Decimal | None = None):
        """Initialize a Position.

        Args:
            symbol: Ticker symbol of the held asset.
            quantity: Number of shares/units held.
            position_datetime: The datetime this position snapshot represents.
            total_value: Market value of the position in the portfolio's primary currency.
            book_value: Cost basis of the position. None if unavailable.
            unit_price: Price per share/unit. None if unavailable.
        """
        self.symbol:str = symbol
        self.quantity:Decimal = quantity
        self.position_datetime:datetime = position_datetime
        self.total_value:Decimal = total_value
        self.book_value:Decimal | None = book_value
        self.unit_price:Decimal | None = unit_price

    @property
    def gain_loss(self) -> Decimal | None:
        """Return the gain/loss (market value - book value)."""
        if self.book_value is None:
            return None
        return self.total_value - self.book_value

    @property
    def gain_loss_percent(self) -> Decimal | None:
        """Return the gain/loss as a percentage of book value."""
        if self.book_value is None or self.book_value == 0:
            return None
        return ((self.total_value - self.book_value) / abs(self.book_value)) * 100

    def __repr__(self):
        return f"Position(ticker={self.symbol}, quantity={self.quantity}, unit_price={self.unit_price})"

class Transaction():
    """A single portfolio transaction (buy, sell, cash movement, etc.)."""

    def __init__(self, symbol:str, transaction_datetime:datetime, transaction_type:TransactionType, quantity:Decimal, price: Union[float, int, Decimal], currency:Currency=Currency.USD):
        """Initialize a Transaction.

        Args:
            symbol: Ticker symbol or target currency code (for CURRENCY_EXCHANGE).
            transaction_datetime: When the transaction occurred (timezone-aware).
            transaction_type: The type of transaction.
            quantity: Number of shares/units transacted.
            price: Price per share/unit.
            currency: Currency of the transaction. Defaults to USD.
        """
        self.symbol:str = symbol
        self.transaction_datetime:datetime = transaction_datetime
        self.transaction_type:TransactionType = transaction_type
        self.quantity:Decimal = quantity
        self.price: Union[float, int, Decimal] = price
        self.currency:Currency = currency

    def __repr__(self):
        return f"Transaction(ticker={self.symbol}, date={self.transaction_datetime}, type={self.transaction_type}, quantity={self.quantity}, price={self.price}, currency={self.currency})"

class Portfolio():
    """A collection of transactions representing an investment portfolio.

    Tracks transactions, cash balances, and holdings with support for
    multi-currency portfolios via an exchange rate manager.
    """

    def __init__(
        self,
        transactions: list[Transaction],
        primary_currency: Currency = Currency.USD,
        error_out_negative_cash: bool = True,
        error_out_negative_quantity: bool = True,
        exchange_rate_manager: ExchangeRateManager | None = None,
        pricing_manager: PricingDataManager | None = None
    ):
        """Initialize a Portfolio.

        Args:
            transactions: List of transactions to include in the portfolio.
            primary_currency: The portfolio's base currency for reporting.
            error_out_negative_cash: If True, raise ValueError when any currency
                balance goes negative during processing.
            error_out_negative_quantity: If True, raise ValueError when any
                holding quantity goes negative during processing.
            exchange_rate_manager: Manager for currency conversions. Defaults to
                FederalReserveExchangeRateManager if not provided.
            pricing_manager: Manager for fetching asset prices. Defaults to
                YFinancePricingDataManager if not provided.
        """
        self.primary_currency = primary_currency
        self.currencies: defaultdict[Currency, Decimal] = defaultdict(Decimal)
        self.transactions: list[Transaction] = transactions

        # If set to True, will raise an error if the cash balance goes negative.
        self.error_out_negative_cash = error_out_negative_cash

        # If set to True, will raise an error if the quantity of any stock/symbol goes negative.
        self.error_out_negative_quantity = error_out_negative_quantity

        # Exchange rate manager for currency conversions
        if exchange_rate_manager:
            self.exchange_rate_manager = exchange_rate_manager
        else:
            self.exchange_rate_manager = FederalReserveExchangeRateManager(
                min_datetime=datetime(2020, 1, 1),
                max_datetime=datetime.now()
            )

        # Pricing data manager for getting asset prices
        if pricing_manager:
            self.pricing_manager = pricing_manager
        else:
            self.pricing_manager = YFinancePricingDataManager()

    def add_transaction_now(
        self,
        symbol: str,
        transaction_type: TransactionType,
        quantity: Decimal,
        price: Union[float, int, Decimal],
        currency: Currency | None = None,
        transaction_datetime: datetime | None = None
    ) -> Transaction:
        """
        Add a transaction to the portfolio.

        If transaction_datetime is None, the current timezone-aware UTC time is used.
        All datetimes are ensured to be timezone-aware.

        Args:
            symbol: Stock symbol for the transaction.
            transaction_type: Type of transaction (BUY, SELL, etc.).
            quantity: Number of shares/units.
            price: Price per share/unit.
            currency: Currency for the transaction. Defaults to portfolio's primary currency.
            transaction_datetime: When the transaction occurred. Defaults to current UTC time.

        Returns:
            The created Transaction object.
        """
        if currency is None:
            currency = self.primary_currency

        if transaction_datetime is None:
            transaction_datetime = datetime.now(timezone.utc)
        elif transaction_datetime.tzinfo is None:
            # Make naive datetime timezone-aware by assuming UTC
            transaction_datetime = transaction_datetime.replace(tzinfo=timezone.utc)

        transaction = Transaction(
            symbol=symbol,
            transaction_datetime=transaction_datetime,
            transaction_type=transaction_type,
            quantity=quantity,
            price=price,
            currency=currency
        )

        self.transactions.append(transaction)
        return transaction

def _create_empty_portfolio_excel(file_path: str) -> None:
    """Create an empty Excel file with just the required headers.

    Args:
        file_path: Path where the Excel file will be created.
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    headers = ["SYMBOL", "DATE AND TIME", "TRANSACTION TYPE", "PRICE", "QUANTITY", "CURRENCY"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    wb.save(file_path)


def load_portfolio_from_excel(
    file_path: str,
    primary_currency: Currency = Currency.USD,
    error_out_negative_cash: bool = True,
    error_out_negative_quantity: bool = True,
    create_if_missing: bool = False,
    pricing_manager: PricingDataManager | None = None,
) -> Portfolio:
    """
    Load a portfolio from an Excel file.

    Args:
        file_path: Path to the Excel file containing transactions.
        primary_currency: Primary currency for the portfolio.
        error_out_negative_cash: If True, raise error on negative cash balance.
        error_out_negative_quantity: If True, raise error on negative quantity.
        create_if_missing: If True and file doesn't exist, create an empty file
                          with headers and return an empty portfolio.
        pricing_manager: Optional pricing data manager for the portfolio.
            Defaults to YFinancePricingDataManager if not provided.

    Returns:
        Portfolio object populated with transactions from the Excel file.

    Expected Excel columns (order independent):
        - SYMBOL: Stock symbol
        - DATE AND TIME: Transaction datetime (ISO format)
        - TRANSACTION TYPE: BUY, SELL, etc.
        - PRICE: Transaction price
        - QUANTITY: Number of shares
        - CURRENCY: Currency code (e.g., CAD, USD). If empty/None for a row,
                   uses the currency from the first row as default.
    """
    import os

    if not os.path.exists(file_path):
        if create_if_missing:
            _create_empty_portfolio_excel(file_path)
            return Portfolio(
                transactions=[],
                primary_currency=primary_currency,
                error_out_negative_cash=error_out_negative_cash,
                error_out_negative_quantity=error_out_negative_quantity,
                pricing_manager=pricing_manager,
            )
        else:
            raise FileNotFoundError(f"Portfolio file not found: {file_path}")

    df = pd.read_excel(file_path)

    if df.empty:
        # File exists with headers but no transactions - return empty portfolio
        return Portfolio(
            transactions=[],
            primary_currency=primary_currency,
            error_out_negative_cash=error_out_negative_cash,
            error_out_negative_quantity=error_out_negative_quantity,
            pricing_manager=pricing_manager,
        )

    required_columns = {"SYMBOL", "DATE AND TIME", "TRANSACTION TYPE", "PRICE", "QUANTITY", "CURRENCY"}
    missing_columns = required_columns - set(df.columns)
    if missing_columns:
        raise ValueError(f"Missing required columns: {missing_columns}")

    # Get default currency from first row
    first_currency_value = df["CURRENCY"].iloc[0]
    default_currency = Currency(first_currency_value)

    transactions: list[Transaction] = []
    any_missing_time = False
    any_missing_timezone = False

    for _, row in df.iterrows():
        symbol = str(row["SYMBOL"])
        raw_datetime = pd.to_datetime(row["DATE AND TIME"]).to_pydatetime() # type: ignore[assignment]
        transaction_datetime, time_missing, tz_missing = _normalize_transaction_datetime(raw_datetime)
        any_missing_time = any_missing_time or time_missing
        any_missing_timezone = any_missing_timezone or tz_missing

        transaction_type = TransactionType(row["TRANSACTION TYPE"])
        price = Decimal(str(row["PRICE"]))
        quantity = Decimal(str(row["QUANTITY"]))

        # Use row currency if present, otherwise use default from first row
        currency_value = row["CURRENCY"]
        if pd.notna(currency_value) and currency_value:
            currency = Currency(currency_value)
        else:
            currency = default_currency

        transaction = Transaction(
            symbol=symbol,
            transaction_datetime=transaction_datetime,
            transaction_type=transaction_type,
            quantity=quantity,
            price=price,
            currency=currency
        )
        transactions.append(transaction)

    # Emit warnings once after processing all transactions
    if any_missing_time and any_missing_timezone:
        warnings.warn(
            f"Some transactions in '{file_path}' were missing time and timezone information. "
            f"Assuming 12:00 PM NYC time (America/New_York) for these transactions.",
            UserWarning
        )
    elif any_missing_time:
        warnings.warn(
            f"Some transactions in '{file_path}' were missing time information. "
            f"Assuming 12:00 PM for these transactions.",
            UserWarning
        )
    elif any_missing_timezone:
        warnings.warn(
            f"Some transactions in '{file_path}' were missing timezone information. "
            f"Assuming NYC timezone (America/New_York) for these transactions.",
            UserWarning
        )

    return Portfolio(
        transactions=transactions,
        primary_currency=primary_currency,
        error_out_negative_cash=error_out_negative_cash,
        error_out_negative_quantity=error_out_negative_quantity,
        pricing_manager=pricing_manager,
    )

def save_portfolio_to_excel(
    portfolio: Portfolio,
    file_path: str
) -> None:
    """
    Save a portfolio to an Excel file.

    Args:
        portfolio: Portfolio object containing transactions to save.
        file_path: Path to the Excel file to write.

    The Excel file will have the following columns:
        - SYMBOL: Stock symbol
        - DATE AND TIME: Transaction datetime (ISO format)
        - TRANSACTION TYPE: BUY, SELL, etc.
        - PRICE: Transaction price
        - QUANTITY: Number of shares
        - CURRENCY: Currency code (e.g., CAD, USD)
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None

    # Write header row
    headers = ["SYMBOL", "DATE AND TIME", "TRANSACTION TYPE", "PRICE", "QUANTITY", "CURRENCY"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Write transaction data
    for row, txn in enumerate(portfolio.transactions, start=2):
        ws.cell(row=row, column=1, value=txn.symbol)
        ws.cell(row=row, column=2, value=txn.transaction_datetime.isoformat())
        ws.cell(row=row, column=3, value=txn.transaction_type.value)
        ws.cell(row=row, column=4, value=float(txn.price))
        ws.cell(row=row, column=5, value=float(txn.quantity))
        ws.cell(row=row, column=6, value=txn.currency.value)

    wb.save(file_path)


def load_portfolio_from_json(
    file_path: str,
    primary_currency: Currency = Currency.USD,
    error_out_negative_cash: bool = True,
    error_out_negative_quantity: bool = True,
    pricing_manager: PricingDataManager | None = None,
) -> Portfolio:
    """
    Load a portfolio from a JSON file.

    Args:
        file_path: Path to the JSON file containing transactions.
        primary_currency: Primary currency for the portfolio.
        error_out_negative_cash: If True, raise error on negative cash balance.
        error_out_negative_quantity: If True, raise error on negative quantity.
        pricing_manager: Optional pricing data manager for the portfolio.
            Defaults to YFinancePricingDataManager if not provided.

    Returns:
        Portfolio object populated with transactions from the JSON file.

    Expected JSON structure:
        [
            {
                "symbol": "AAPL",
                "datetime": "2024-01-15T10:30:00",
                "transaction_type": "BUY",
                "price": 150.50,
                "quantity": 10,
                "currency": "USD"
            },
            ...
        ]
    """
    with open(file_path, "r") as f:
        data = json.load(f)

    if not isinstance(data, list):
        raise ValueError("JSON file must contain a list of transactions")

    transactions: list[Transaction] = []
    any_missing_time = False
    any_missing_timezone = False

    item: Any
    for item in data:  # type: ignore[union-attr]
        symbol = str(item["symbol"])
        raw_datetime = datetime.fromisoformat(item["datetime"])
        transaction_datetime, time_missing, tz_missing = _normalize_transaction_datetime(raw_datetime)
        any_missing_time = any_missing_time or time_missing
        any_missing_timezone = any_missing_timezone or tz_missing

        transaction_type = TransactionType(item["transaction_type"])
        price = Decimal(str(item["price"]))
        quantity = Decimal(str(item["quantity"]))
        currency = Currency(item["currency"])

        transaction = Transaction(
            symbol=symbol,
            transaction_datetime=transaction_datetime,
            transaction_type=transaction_type,
            quantity=quantity,
            price=price,
            currency=currency
        )
        transactions.append(transaction)

    # Emit warnings once after processing all transactions
    if any_missing_time and any_missing_timezone:
        warnings.warn(
            f"Some transactions in '{file_path}' were missing time and timezone information. "
            f"Assuming 12:00 PM NYC time (America/New_York) for these transactions.",
            UserWarning
        )
    elif any_missing_time:
        warnings.warn(
            f"Some transactions in '{file_path}' were missing time information. "
            f"Assuming 12:00 PM for these transactions.",
            UserWarning
        )
    elif any_missing_timezone:
        warnings.warn(
            f"Some transactions in '{file_path}' were missing timezone information. "
            f"Assuming NYC timezone (America/New_York) for these transactions.",
            UserWarning
        )

    return Portfolio(
        transactions=transactions,
        primary_currency=primary_currency,
        error_out_negative_cash=error_out_negative_cash,
        error_out_negative_quantity=error_out_negative_quantity,
        pricing_manager=pricing_manager,
    )


def save_portfolio_to_json(
    portfolio: Portfolio,
    file_path: str
) -> None:
    """
    Save a portfolio to a JSON file.

    Args:
        portfolio: Portfolio object containing transactions to save.
        file_path: Path to the JSON file to write.

    The JSON file will contain a list of transactions with the following fields:
        - symbol: Stock symbol
        - datetime: Transaction datetime (ISO format)
        - transaction_type: BUY, SELL, etc.
        - price: Transaction price
        - quantity: Number of shares
        - currency: Currency code (e.g., CAD, USD)
    """
    data = []
    for txn in portfolio.transactions:
        data.append({
            "symbol": txn.symbol,
            "datetime": txn.transaction_datetime.isoformat(),
            "transaction_type": txn.transaction_type.value,
            "price": float(txn.price),
            "quantity": float(txn.quantity),
            "currency": txn.currency.value
        })

    with open(file_path, "w") as f:
        json.dump(data, f, indent=2)


def get_cash_balances(
    portfolio: Portfolio,
    as_of: datetime | None = None
) -> dict[Currency, Decimal]:
    """
    Get cash balances by currency as of a specific datetime.

    Args:
        portfolio: Portfolio object containing transactions.
        as_of: The datetime to calculate balances for. If None, uses all transactions.

    Returns:
        A dictionary mapping Currency to Decimal balance.
        Only includes currencies with non-zero balances.
    """
    currencies: dict[Currency, Decimal] = defaultdict(Decimal)

    sorted_transactions = sorted(
        portfolio.transactions,
        key=lambda t: t.transaction_datetime
    )

    for txn in sorted_transactions:
        if as_of is not None and txn.transaction_datetime > as_of:
            break

        currency = txn.currency
        quantity = txn.quantity
        price = Decimal(str(txn.price))
        total_value = quantity * price

        if txn.transaction_type == TransactionType.BUY:
            currencies[currency] -= total_value

        elif txn.transaction_type == TransactionType.SELL:
            currencies[currency] += total_value

        elif txn.transaction_type == TransactionType.CASH_IN:
            currencies[currency] += total_value

        elif txn.transaction_type == TransactionType.CASH_OUT:
            currencies[currency] -= total_value

        elif txn.transaction_type == TransactionType.DIVIDEND:
            currencies[currency] += total_value

        elif txn.transaction_type == TransactionType.INTEREST:
            currencies[currency] += total_value

        elif txn.transaction_type == TransactionType.FEE:
            currencies[currency] -= total_value

        elif txn.transaction_type == TransactionType.CURRENCY_EXCHANGE:
            # symbol is target currency, currency is source currency
            # price is cost in source currency per 1 unit of target
            # quantity is amount of target currency purchased
            target_currency = Currency(txn.symbol)
            currencies[currency] -= total_value  # spend source currency
            currencies[target_currency] += quantity  # receive target currency

    # Filter out zero balances
    return {curr: bal for curr, bal in currencies.items() if bal != Decimal("0")}


def calculate_portfolio_final_cash_available(
    portfolio: Portfolio,
    target_currency: Currency
) -> Decimal:
    """
    Calculate the final portfolio cash value by processing all transactions chronologically.

    Args:
        portfolio: Portfolio object containing transactions and settings.
        target_currency: The currency to convert all values into.

    Returns:
        A Decimal representing the total cash value in the target currency.

    Raises:
        ValueError: If error_out_negative_cash is True and any currency goes negative.
        ValueError: If error_out_negative_quantity is True and any holding goes negative.
    """
    currencies: dict[Currency, Decimal] = defaultdict(Decimal)
    holdings: dict[str, Decimal] = defaultdict(Decimal)

    sorted_transactions = sorted(
        portfolio.transactions,
        key=lambda t: t.transaction_datetime
    )

    for txn in sorted_transactions:
        symbol = txn.symbol
        currency = txn.currency
        quantity = txn.quantity
        price = Decimal(str(txn.price))
        total_value = quantity * price

        if txn.transaction_type == TransactionType.BUY:
            currencies[currency] -= total_value
            holdings[symbol] += quantity

        elif txn.transaction_type == TransactionType.SELL:
            currencies[currency] += total_value
            holdings[symbol] -= quantity

        elif txn.transaction_type == TransactionType.CASH_IN:
            currencies[currency] += total_value

        elif txn.transaction_type == TransactionType.CASH_OUT:
            currencies[currency] -= total_value

        elif txn.transaction_type == TransactionType.DIVIDEND:
            currencies[currency] += total_value

        elif txn.transaction_type == TransactionType.INTEREST:
            currencies[currency] += total_value

        elif txn.transaction_type == TransactionType.FEE:
            currencies[currency] -= total_value

        elif txn.transaction_type == TransactionType.CURRENCY_EXCHANGE:
            # symbol is target currency, currency is source currency
            # price is cost in source currency per 1 unit of target
            # quantity is amount of target currency purchased
            exchange_target_currency = Currency(symbol)
            currencies[currency] -= total_value  # spend source currency
            currencies[exchange_target_currency] += quantity  # receive target currency

        if portfolio.error_out_negative_cash:
            for curr, balance in currencies.items():
                if balance < 0:
                    raise ValueError(
                        f"Negative cash balance detected: {curr.value} = {balance} "
                        f"after transaction: {txn}"
                    )

        if portfolio.error_out_negative_quantity:
            for sym, qty in holdings.items():
                if qty < 0:
                    raise ValueError(
                        f"Negative quantity detected: {sym} = {qty} "
                        f"after transaction: {txn}"
                    )

    total_in_target_currency = Decimal("0")
    for currency, balance in currencies.items():
        rate = portfolio.exchange_rate_manager.get_exchange_rate(currency, target_currency)
        total_in_target_currency += balance * rate

    return total_in_target_currency

def convert_currencies_to_target(
    currencies: dict[Currency, Decimal],
    target_currency: Currency,
    exchange_rate_manager: ExchangeRateManager,
    date: datetime
) -> Decimal:
    """
    Convert a dictionary of currency balances to a single target currency value.

    Args:
        currencies: Dictionary mapping Currency to Decimal balance.
        target_currency: The currency to convert all balances into.
        exchange_rate_manager: The exchange rate manager to use for conversions.
        date: The date to use for exchange rate lookups.

    Returns:
        A Decimal representing the total value in the target currency.
    """
    total = Decimal("0")
    for currency, balance in currencies.items():
        rate = exchange_rate_manager.get_exchange_rate(currency, target_currency, date)
        total += balance * rate
    return total


def calculate_cash_available_by_day(
    portfolio: Portfolio,
    target_currency: Currency,
    start_date: datetime | None = None,
    end_date: datetime | None = None
) -> dict[datetime, Decimal]:
    """
    Calculate the portfolio cash value for each day within a date range.

    Args:
        portfolio: Portfolio object containing transactions and settings.
        target_currency: The currency to convert all values into.
        start_date: The start date for the calculation (inclusive).
                    If None, uses the earliest transaction date.
        end_date: The end date for the calculation (inclusive).
                  If None, uses the latest transaction date.

    Returns:
        A dictionary mapping each date to the portfolio value in the target currency.

    Raises:
        ValueError: If error_out_negative_cash is True and any currency goes negative.
        ValueError: If error_out_negative_quantity is True and any holding goes negative.
    """
    if not portfolio.transactions:
        return {}

    sorted_transactions = sorted(
        portfolio.transactions,
        key=lambda t: t.transaction_datetime
    )

    # Determine date range
    if start_date is None:
        start_date = sorted_transactions[0].transaction_datetime
    if end_date is None:
        end_date = datetime.now()

    # Normalize to date boundaries (start of day)
    start_date = datetime(start_date.year, start_date.month, start_date.day)
    end_date = datetime(end_date.year, end_date.month, end_date.day)

    currencies: dict[Currency, Decimal] = defaultdict(Decimal)
    holdings: dict[str, Decimal] = defaultdict(Decimal)
    result: dict[datetime, Decimal] = {}

    # Create an iterator over transactions
    txn_index = 0

    # Iterate through each day in the range
    current_date = start_date
    while current_date <= end_date:
        # Process all transactions that occurred on or before current_date
        while txn_index < len(sorted_transactions):
            txn = sorted_transactions[txn_index]
            txn_date = datetime(
                txn.transaction_datetime.year,
                txn.transaction_datetime.month,
                txn.transaction_datetime.day
            )

            if txn_date > current_date:
                break

            # Process the transaction
            symbol = txn.symbol
            currency = txn.currency
            quantity = txn.quantity
            price = Decimal(str(txn.price))
            total_value = quantity * price

            if txn.transaction_type == TransactionType.BUY:
                currencies[currency] -= total_value
                holdings[symbol] += quantity

            elif txn.transaction_type == TransactionType.SELL:
                currencies[currency] += total_value
                holdings[symbol] -= quantity

            elif txn.transaction_type == TransactionType.CASH_IN:
                currencies[currency] += total_value

            elif txn.transaction_type == TransactionType.CASH_OUT:
                currencies[currency] -= total_value

            elif txn.transaction_type == TransactionType.DIVIDEND:
                currencies[currency] += total_value

            elif txn.transaction_type == TransactionType.INTEREST:
                currencies[currency] += total_value

            elif txn.transaction_type == TransactionType.FEE:
                currencies[currency] -= total_value

            elif txn.transaction_type == TransactionType.CURRENCY_EXCHANGE:
                # symbol is target currency, currency is source currency
                # price is cost in source currency per 1 unit of target
                # quantity is amount of target currency purchased
                exchange_target_currency = Currency(symbol)
                currencies[currency] -= total_value  # spend source currency
                currencies[exchange_target_currency] += quantity  # receive target currency

            if portfolio.error_out_negative_cash:
                for curr, balance in currencies.items():
                    if balance < 0:
                        raise ValueError(
                            f"Negative cash balance detected: {curr.value} = {balance} "
                            f"after transaction: {txn}"
                        )

            if portfolio.error_out_negative_quantity:
                for sym, qty in holdings.items():
                    if qty < 0:
                        raise ValueError(
                            f"Negative quantity detected: {sym} = {qty} "
                            f"after transaction: {txn}"
                        )

            txn_index += 1

        # Calculate the portfolio value for this day
        daily_value = convert_currencies_to_target(
            currencies,
            target_currency,
            portfolio.exchange_rate_manager,
            current_date
        )
        result[current_date] = daily_value

        # Move to next day
        current_date = current_date + timedelta(days=1)

    return result


def calculate_portfolio_book_value_on_date(
    portfolio: Portfolio,
    valuation_date: datetime,
    target_currency: Currency
) -> Decimal:
    """
    Calculate the total portfolio book value (cash + holdings at cost) on a specific date.

    This function processes all transactions up to the valuation date to determine
    holdings and their cost basis, then converts everything to the target currency.
    Unlike calculate_portfolio_value_on_date, this uses the purchase price (book value)
    rather than current market prices.

    Args:
        portfolio: Portfolio object containing transactions and settings.
        valuation_date: The date to calculate the portfolio book value for.
        target_currency: The currency to convert all values into.

    Returns:
        A Decimal representing the total portfolio book value in the target currency.

    Raises:
        ValueError: If error_out_negative_cash is True and any currency goes negative.
        ValueError: If error_out_negative_quantity is True and any holding goes negative.
    """
    currencies: dict[Currency, Decimal] = defaultdict(Decimal)
    holdings: dict[str, Decimal] = defaultdict(Decimal)
    holding_costs: dict[str, Decimal] = defaultdict(Decimal)  # Track total cost basis for each holding
    holding_currencies: dict[str, Currency] = {}  # Track the currency each holding was bought in

    # Process all transactions up to and including the valuation date
    sorted_transactions = sorted(
        portfolio.transactions,
        key=lambda t: t.transaction_datetime
    )

    for txn in sorted_transactions:
        # Skip transactions after the valuation date
        if txn.transaction_datetime > valuation_date:
            break

        symbol = txn.symbol
        currency = txn.currency
        quantity = txn.quantity
        price = Decimal(str(txn.price))
        total_value = quantity * price

        if txn.transaction_type == TransactionType.BUY:
            currencies[currency] -= total_value
            if holdings[symbol] < 0:
                # Covering a short: reduce short cost basis proportionally
                cover_quantity = min(quantity, abs(holdings[symbol]))
                cost_per_unit = holding_costs[symbol] / holdings[symbol]
                holding_costs[symbol] += cost_per_unit * cover_quantity
                long_quantity = quantity - cover_quantity
                if long_quantity > 0:
                    holding_costs[symbol] += long_quantity * price
            else:
                holding_costs[symbol] += total_value
            holding_currencies[symbol] = currency
            holdings[symbol] += quantity

        elif txn.transaction_type == TransactionType.SELL:
            currencies[currency] += total_value
            if holdings[symbol] > 0:
                sell_from_long = min(quantity, holdings[symbol])
                cost_per_unit = holding_costs[symbol] / holdings[symbol]
                holding_costs[symbol] -= cost_per_unit * sell_from_long
                short_quantity = quantity - sell_from_long
                if short_quantity > 0:
                    # Remaining goes short: record short cost basis (negative)
                    holding_costs[symbol] -= short_quantity * price
            elif holdings[symbol] < 0:
                # Deepen existing short
                holding_costs[symbol] -= total_value
            else:
                # Holdings == 0: open new short
                holding_costs[symbol] = -total_value
            holding_currencies[symbol] = currency
            holdings[symbol] -= quantity

        elif txn.transaction_type == TransactionType.CASH_IN:
            currencies[currency] += total_value

        elif txn.transaction_type == TransactionType.CASH_OUT:
            currencies[currency] -= total_value

        elif txn.transaction_type == TransactionType.DIVIDEND:
            currencies[currency] += total_value

        elif txn.transaction_type == TransactionType.INTEREST:
            currencies[currency] += total_value

        elif txn.transaction_type == TransactionType.FEE:
            currencies[currency] -= total_value

        elif txn.transaction_type == TransactionType.CURRENCY_EXCHANGE:
            # symbol is target currency, currency is source currency
            # price is cost in source currency per 1 unit of target
            # quantity is amount of target currency purchased
            exchange_target_currency = Currency(symbol)
            currencies[currency] -= total_value  # spend source currency
            currencies[exchange_target_currency] += quantity  # receive target currency

        if portfolio.error_out_negative_cash:
            for curr, balance in currencies.items():
                if balance < 0:
                    raise ValueError(
                        f"Negative cash balance detected: {curr.value} = {balance} "
                        f"after transaction: {txn}"
                    )

        if portfolio.error_out_negative_quantity:
            for sym, qty in holdings.items():
                if qty < 0:
                    raise ValueError(
                        f"Negative quantity detected: {sym} = {qty} "
                        f"after transaction: {txn}"
                    )

    # Calculate total cash value in target currency
    total_cash = convert_currencies_to_target(
        currencies,
        target_currency,
        portfolio.exchange_rate_manager,
        valuation_date
    )

    # Calculate total holdings book value in target currency
    total_holdings_book_value = Decimal("0")
    for symbol, quantity in holdings.items():
        if quantity != 0:
            # Use the cost basis instead of market price
            holding_book_value = holding_costs[symbol]

            # Use the currency from the BUY transaction
            holding_base_currency = holding_currencies.get(symbol, Currency.USD)

            # Convert to target currency if needed
            if holding_base_currency != target_currency:
                rate = portfolio.exchange_rate_manager.get_exchange_rate(
                    holding_base_currency,
                    target_currency,
                    valuation_date
                )
                holding_book_value = holding_book_value * rate

            total_holdings_book_value += holding_book_value

    return total_cash + total_holdings_book_value


def calculate_portfolio_value_on_date(
    portfolio: Portfolio,
    valuation_date: datetime,
    target_currency: Currency
) -> Decimal:
    """
    Calculate the total portfolio value (cash + holdings) on a specific date.

    This function processes all transactions up to the valuation date to determine
    holdings, then uses the portfolio's pricing manager to get current prices
    for each holding and converts everything to the target currency.

    Args:
        portfolio: Portfolio object containing transactions and settings.
        valuation_date: The date to calculate the portfolio value for.
        target_currency: The currency to convert all values into.

    Returns:
        A Decimal representing the total portfolio value in the target currency.

    Raises:
        ValueError: If error_out_negative_cash is True and any currency goes negative.
        ValueError: If error_out_negative_quantity is True and any holding goes negative.
    """
    currencies: dict[Currency, Decimal] = defaultdict(Decimal)
    holdings: dict[str, Decimal] = defaultdict(Decimal)
    holding_currencies: dict[str, Currency] = {}  # Track the currency each holding was bought in

    # Process all transactions up to and including the valuation date
    sorted_transactions = sorted(
        portfolio.transactions,
        key=lambda t: t.transaction_datetime
    )

    for txn in sorted_transactions:
        # Skip transactions after the valuation date
        if txn.transaction_datetime > valuation_date:
            break

        symbol = txn.symbol
        currency = txn.currency
        quantity = txn.quantity
        price = Decimal(str(txn.price))
        total_value = quantity * price

        if txn.transaction_type == TransactionType.BUY:
            currencies[currency] -= total_value
            holdings[symbol] += quantity
            holding_currencies[symbol] = currency  # Record the currency used for this holding

        elif txn.transaction_type == TransactionType.SELL:
            currencies[currency] += total_value
            holdings[symbol] -= quantity

        elif txn.transaction_type == TransactionType.CASH_IN:
            currencies[currency] += total_value

        elif txn.transaction_type == TransactionType.CASH_OUT:
            currencies[currency] -= total_value

        elif txn.transaction_type == TransactionType.DIVIDEND:
            currencies[currency] += total_value

        elif txn.transaction_type == TransactionType.INTEREST:
            currencies[currency] += total_value

        elif txn.transaction_type == TransactionType.FEE:
            currencies[currency] -= total_value

        elif txn.transaction_type == TransactionType.CURRENCY_EXCHANGE:
            # symbol is target currency, currency is source currency
            # price is cost in source currency per 1 unit of target
            # quantity is amount of target currency purchased
            exchange_target_currency = Currency(symbol)
            currencies[currency] -= total_value  # spend source currency
            currencies[exchange_target_currency] += quantity  # receive target currency

        if portfolio.error_out_negative_cash:
            for curr, balance in currencies.items():
                if balance < 0:
                    raise ValueError(
                        f"Negative cash balance detected: {curr.value} = {balance} "
                        f"after transaction: {txn}"
                    )

        if portfolio.error_out_negative_quantity:
            for sym, qty in holdings.items():
                if qty < 0:
                    raise ValueError(
                        f"Negative quantity detected: {sym} = {qty} "
                        f"after transaction: {txn}"
                    )

    # Calculate total cash value in target currency
    total_cash = convert_currencies_to_target(
        currencies,
        target_currency,
        portfolio.exchange_rate_manager,
        valuation_date
    )

    # Calculate total holdings value in target currency
    total_holdings_value = Decimal("0")
    for symbol, quantity in holdings.items():
        if quantity != 0:
            # Get the price for this symbol on the valuation date
            price_point = portfolio.pricing_manager.get_price_point(symbol, valuation_date)

            # Calculate the value in the holding's base currency (from BUY transaction)
            holding_value = quantity * price_point.price

            # Use the currency from the BUY transaction, not the pricing manager
            holding_base_currency = holding_currencies.get(symbol, Currency.USD)

            # Convert to target currency if needed
            if holding_base_currency != target_currency:
                rate = portfolio.exchange_rate_manager.get_exchange_rate(
                    holding_base_currency,
                    target_currency,
                    valuation_date
                )
                holding_value = holding_value * rate

            total_holdings_value += holding_value

    return total_cash + total_holdings_value


def calculate_portfolio_value_by_day(
    portfolio: Portfolio,
    target_currency: Currency,
    start_date: datetime | None = None,
    end_date: datetime | None = None
) -> dict[datetime, Decimal]:
    """
    Calculate the total portfolio value (cash + holdings) for each day within a date range.

    This function processes all transactions chronologically and calculates the
    portfolio value (cash + holdings valued at market prices) for each day in the range.

    Args:
        portfolio: Portfolio object containing transactions and settings.
        target_currency: The currency to convert all values into.
        start_date: The start date for the calculation (inclusive).
                    If None, uses the earliest transaction date.
        end_date: The end date for the calculation (inclusive).
                  If None, uses the latest transaction date.

    Returns:
        A dictionary mapping each date to the total portfolio value in the target currency.

    Raises:
        ValueError: If error_out_negative_cash is True and any currency goes negative.
        ValueError: If error_out_negative_quantity is True and any holding goes negative.
    """
    if not portfolio.transactions:
        return {}

    sorted_transactions = sorted(
        portfolio.transactions,
        key=lambda t: t.transaction_datetime
    )

    # Determine date range
    if start_date is None:
        start_date = sorted_transactions[0].transaction_datetime
    if end_date is None:
        end_date = datetime.now()

    # Normalize to date boundaries (start of day)
    start_date = datetime(start_date.year, start_date.month, start_date.day)
    end_date = datetime(end_date.year, end_date.month, end_date.day)

    currencies: dict[Currency, Decimal] = defaultdict(Decimal)
    holdings: dict[str, Decimal] = defaultdict(Decimal)
    holding_currencies: dict[str, Currency] = {}  # Track the currency each holding was bought in
    result: dict[datetime, Decimal] = {}

    # Create an iterator over transactions
    txn_index = 0

    # Iterate through each day in the range
    current_date = start_date
    while current_date <= end_date:
        # Process all transactions that occurred on or before current_date
        while txn_index < len(sorted_transactions):
            txn = sorted_transactions[txn_index]
            txn_date = datetime(
                txn.transaction_datetime.year,
                txn.transaction_datetime.month,
                txn.transaction_datetime.day
            )

            if txn_date > current_date:
                break

            # Process the transaction
            symbol = txn.symbol
            currency = txn.currency
            quantity = txn.quantity
            price = Decimal(str(txn.price))
            total_value = quantity * price

            if txn.transaction_type == TransactionType.BUY:
                currencies[currency] -= total_value
                holdings[symbol] += quantity
                holding_currencies[symbol] = currency  # Record the currency used for this holding

            elif txn.transaction_type == TransactionType.SELL:
                currencies[currency] += total_value
                holdings[symbol] -= quantity

            elif txn.transaction_type == TransactionType.CASH_IN:
                currencies[currency] += total_value

            elif txn.transaction_type == TransactionType.CASH_OUT:
                currencies[currency] -= total_value

            elif txn.transaction_type == TransactionType.DIVIDEND:
                currencies[currency] += total_value

            elif txn.transaction_type == TransactionType.INTEREST:
                currencies[currency] += total_value

            elif txn.transaction_type == TransactionType.FEE:
                currencies[currency] -= total_value

            elif txn.transaction_type == TransactionType.CURRENCY_EXCHANGE:
                # symbol is target currency, currency is source currency
                # price is cost in source currency per 1 unit of target
                # quantity is amount of target currency purchased
                exchange_target_currency = Currency(symbol)
                currencies[currency] -= total_value  # spend source currency
                currencies[exchange_target_currency] += quantity  # receive target currency

            if portfolio.error_out_negative_cash:
                for curr, balance in currencies.items():
                    if balance < 0:
                        raise ValueError(
                            f"Negative cash balance detected: {curr.value} = {balance} "
                            f"after transaction: {txn}"
                        )

            if portfolio.error_out_negative_quantity:
                for sym, qty in holdings.items():
                    if qty < 0:
                        raise ValueError(
                            f"Negative quantity detected: {sym} = {qty} "
                            f"after transaction: {txn}"
                        )

            txn_index += 1

        # Calculate the cash value for this day
        daily_cash = convert_currencies_to_target(
            currencies,
            target_currency,
            portfolio.exchange_rate_manager,
            current_date
        )

        # Calculate total holdings value in target currency
        daily_holdings_value = Decimal("0")
        for symbol, quantity in holdings.items():
            if quantity != 0:
                # Get the price for this symbol on the current date
                price_point = portfolio.pricing_manager.get_price_point(symbol, current_date)

                # Calculate the value in the holding's base currency (from BUY transaction)
                holding_value = quantity * price_point.price

                # Use the currency from the BUY transaction, not the pricing manager
                holding_base_currency = holding_currencies.get(symbol, Currency.USD)

                # Convert to target currency if needed
                if holding_base_currency != target_currency:
                    rate = portfolio.exchange_rate_manager.get_exchange_rate(
                        holding_base_currency,
                        target_currency,
                        current_date
                    )
                    holding_value = holding_value * rate

                daily_holdings_value += holding_value

        result[current_date] = daily_cash + daily_holdings_value

        # Move to next day
        current_date = current_date + timedelta(days=1)

    return result


def calculate_portfolio_book_value_by_day(
    portfolio: Portfolio,
    target_currency: Currency,
    start_date: datetime | None = None,
    end_date: datetime | None = None
) -> dict[datetime, Decimal]:
    """
    Calculate the total portfolio book value (cash + holdings at cost) for each day within a date range.

    This function processes all transactions chronologically and calculates the
    portfolio book value (cash + holdings valued at cost basis) for each day in the range.
    Unlike calculate_portfolio_value_by_day, this uses the purchase price (book value)
    rather than current market prices.

    Args:
        portfolio: Portfolio object containing transactions and settings.
        target_currency: The currency to convert all values into.
        start_date: The start date for the calculation (inclusive).
                    If None, uses the earliest transaction date.
        end_date: The end date for the calculation (inclusive).
                  If None, uses the latest transaction date.

    Returns:
        A dictionary mapping each date to the total portfolio book value in the target currency.

    Raises:
        ValueError: If error_out_negative_cash is True and any currency goes negative.
        ValueError: If error_out_negative_quantity is True and any holding goes negative.
    """
    if not portfolio.transactions:
        return {}

    sorted_transactions = sorted(
        portfolio.transactions,
        key=lambda t: t.transaction_datetime
    )

    # Determine date range
    if start_date is None:
        start_date = sorted_transactions[0].transaction_datetime
    if end_date is None:
        end_date = datetime.now()

    # Normalize to date boundaries (start of day)
    start_date = datetime(start_date.year, start_date.month, start_date.day)
    end_date = datetime(end_date.year, end_date.month, end_date.day)

    currencies: dict[Currency, Decimal] = defaultdict(Decimal)
    holdings: dict[str, Decimal] = defaultdict(Decimal)
    holding_costs: dict[str, Decimal] = defaultdict(Decimal)  # Track total cost basis for each holding
    holding_currencies: dict[str, Currency] = {}  # Track the currency each holding was bought in
    result: dict[datetime, Decimal] = {}

    # Create an iterator over transactions
    txn_index = 0

    # Iterate through each day in the range
    current_date = start_date
    while current_date <= end_date:
        # Process all transactions that occurred on or before current_date
        while txn_index < len(sorted_transactions):
            txn = sorted_transactions[txn_index]
            txn_date = datetime(
                txn.transaction_datetime.year,
                txn.transaction_datetime.month,
                txn.transaction_datetime.day
            )

            if txn_date > current_date:
                break

            # Process the transaction
            symbol = txn.symbol
            currency = txn.currency
            quantity = txn.quantity
            price = Decimal(str(txn.price))
            total_value = quantity * price

            if txn.transaction_type == TransactionType.BUY:
                currencies[currency] -= total_value
                if holdings[symbol] < 0:
                    # Covering a short: reduce short cost basis proportionally
                    cover_quantity = min(quantity, abs(holdings[symbol]))
                    cost_per_unit = holding_costs[symbol] / holdings[symbol]
                    holding_costs[symbol] += cost_per_unit * cover_quantity
                    long_quantity = quantity - cover_quantity
                    if long_quantity > 0:
                        holding_costs[symbol] += long_quantity * price
                else:
                    holding_costs[symbol] += total_value
                holding_currencies[symbol] = currency
                holdings[symbol] += quantity

            elif txn.transaction_type == TransactionType.SELL:
                currencies[currency] += total_value
                if holdings[symbol] > 0:
                    sell_from_long = min(quantity, holdings[symbol])
                    cost_per_unit = holding_costs[symbol] / holdings[symbol]
                    holding_costs[symbol] -= cost_per_unit * sell_from_long
                    short_quantity = quantity - sell_from_long
                    if short_quantity > 0:
                        # Remaining goes short: record short cost basis (negative)
                        holding_costs[symbol] -= short_quantity * price
                elif holdings[symbol] < 0:
                    # Deepen existing short
                    holding_costs[symbol] -= total_value
                else:
                    # Holdings == 0: open new short
                    holding_costs[symbol] = -total_value
                holding_currencies[symbol] = currency
                holdings[symbol] -= quantity

            elif txn.transaction_type == TransactionType.CASH_IN:
                currencies[currency] += total_value

            elif txn.transaction_type == TransactionType.CASH_OUT:
                currencies[currency] -= total_value

            elif txn.transaction_type == TransactionType.DIVIDEND:
                currencies[currency] += total_value

            elif txn.transaction_type == TransactionType.INTEREST:
                currencies[currency] += total_value

            elif txn.transaction_type == TransactionType.FEE:
                currencies[currency] -= total_value

            elif txn.transaction_type == TransactionType.CURRENCY_EXCHANGE:
                # symbol is target currency, currency is source currency
                # price is cost in source currency per 1 unit of target
                # quantity is amount of target currency purchased
                exchange_target_currency = Currency(symbol)
                currencies[currency] -= total_value  # spend source currency
                currencies[exchange_target_currency] += quantity  # receive target currency

            if portfolio.error_out_negative_cash:
                for curr, balance in currencies.items():
                    if balance < 0:
                        raise ValueError(
                            f"Negative cash balance detected: {curr.value} = {balance} "
                            f"after transaction: {txn}"
                        )

            if portfolio.error_out_negative_quantity:
                for sym, qty in holdings.items():
                    if qty < 0:
                        raise ValueError(
                            f"Negative quantity detected: {sym} = {qty} "
                            f"after transaction: {txn}"
                        )

            txn_index += 1

        # Calculate the cash value for this day
        daily_cash = convert_currencies_to_target(
            currencies,
            target_currency,
            portfolio.exchange_rate_manager,
            current_date
        )

        # Calculate total holdings book value in target currency
        daily_holdings_book_value = Decimal("0")
        for symbol, quantity in holdings.items():
            if quantity != 0:
                # Use the cost basis instead of market price
                holding_book_value = holding_costs[symbol]

                # Use the currency from the BUY transaction
                holding_base_currency = holding_currencies.get(symbol, Currency.USD)

                # Convert to target currency if needed
                if holding_base_currency != target_currency:
                    rate = portfolio.exchange_rate_manager.get_exchange_rate(
                        holding_base_currency,
                        target_currency,
                        current_date
                    )
                    holding_book_value = holding_book_value * rate

                daily_holdings_book_value += holding_book_value

        result[current_date] = daily_cash + daily_holdings_book_value

        # Move to next day
        current_date = current_date + timedelta(days=1)

    return result


def get_positions(
    position_datetime: datetime,
    portfolio: Portfolio
) -> list[Position]:
    """
    Get all actively held positions at a specific point in time.

    This function processes all transactions up to the specified datetime
    to determine the quantity of each symbol held, then creates Position
    objects with market values calculated using the pricing manager.

    Args:
        position_datetime: The datetime to calculate positions for.
        portfolio: Portfolio object containing transactions and settings.

    Returns:
        A list of Position objects for all actively held symbols.
        The total_value of each position is calculated using market prices
        and converted to the portfolio's primary currency.

    Raises:
        ValueError: If error_out_negative_cash is True and any currency goes negative.
        ValueError: If error_out_negative_quantity is True and any holding goes negative.
    """
    holdings: dict[str, Decimal] = defaultdict(Decimal)
    holding_currencies: dict[str, Currency] = {}  # Track the currency each holding was bought in
    holding_costs: dict[str, Decimal] = defaultdict(Decimal)  # Track total cost basis for each holding
    currencies: dict[Currency, Decimal] = defaultdict(Decimal)

    sorted_transactions = sorted(
        portfolio.transactions,
        key=lambda t: t.transaction_datetime
    )

    for txn in sorted_transactions:
        if txn.transaction_datetime > position_datetime:
            break

        symbol = txn.symbol
        currency = txn.currency
        quantity = txn.quantity
        price = Decimal(str(txn.price))
        total_value = quantity * price

        if txn.transaction_type == TransactionType.BUY:
            currencies[currency] -= total_value
            if holdings[symbol] < 0:
                # Covering a short: reduce short cost basis proportionally
                cover_quantity = min(quantity, abs(holdings[symbol]))
                cost_per_unit = holding_costs[symbol] / holdings[symbol]
                holding_costs[symbol] += cost_per_unit * cover_quantity
                long_quantity = quantity - cover_quantity
                if long_quantity > 0:
                    holding_costs[symbol] += long_quantity * price
            else:
                holding_costs[symbol] += total_value
            holding_currencies[symbol] = currency
            holdings[symbol] += quantity

        elif txn.transaction_type == TransactionType.SELL:
            currencies[currency] += total_value
            if holdings[symbol] > 0:
                sell_from_long = min(quantity, holdings[symbol])
                cost_per_unit = holding_costs[symbol] / holdings[symbol]
                holding_costs[symbol] -= cost_per_unit * sell_from_long
                short_quantity = quantity - sell_from_long
                if short_quantity > 0:
                    # Remaining goes short: record short cost basis (negative)
                    holding_costs[symbol] -= short_quantity * price
            elif holdings[symbol] < 0:
                # Deepen existing short
                holding_costs[symbol] -= total_value
            else:
                # Holdings == 0: open new short
                holding_costs[symbol] = -total_value
            holding_currencies[symbol] = currency
            holdings[symbol] -= quantity

        elif txn.transaction_type == TransactionType.CASH_IN:
            currencies[currency] += total_value

        elif txn.transaction_type == TransactionType.CASH_OUT:
            currencies[currency] -= total_value

        elif txn.transaction_type == TransactionType.DIVIDEND:
            currencies[currency] += total_value

        elif txn.transaction_type == TransactionType.INTEREST:
            currencies[currency] += total_value

        elif txn.transaction_type == TransactionType.FEE:
            currencies[currency] -= total_value

        elif txn.transaction_type == TransactionType.CURRENCY_EXCHANGE:
            # symbol is target currency, currency is source currency
            # price is cost in source currency per 1 unit of target
            # quantity is amount of target currency purchased
            exchange_target_currency = Currency(symbol)
            currencies[currency] -= total_value  # spend source currency
            currencies[exchange_target_currency] += quantity  # receive target currency

        if portfolio.error_out_negative_cash:
            for curr, balance in currencies.items():
                if balance < 0:
                    raise ValueError(
                        f"Negative cash balance detected: {curr.value} = {balance} "
                        f"after transaction: {txn}"
                    )

        if portfolio.error_out_negative_quantity:
            for sym, qty in holdings.items():
                if qty < 0:
                    raise ValueError(
                        f"Negative quantity detected: {sym} = {qty} "
                        f"after transaction: {txn}"
                    )

    positions: list[Position] = []
    target_currency = portfolio.primary_currency

    for symbol, quantity in holdings.items():
        if quantity != 0:
            price_point = portfolio.pricing_manager.get_price_point(symbol, position_datetime)
            unit_price = price_point.price
            market_value = quantity * unit_price

            # Get the book value (cost basis) for this holding
            book_value = holding_costs[symbol]

            # Use the currency from the BUY transaction, not the pricing manager
            holding_base_currency = holding_currencies.get(symbol, Currency.USD)

            if holding_base_currency != target_currency:
                rate = portfolio.exchange_rate_manager.get_exchange_rate(
                    holding_base_currency,
                    target_currency,
                    position_datetime
                )
                market_value = market_value * rate
                book_value = book_value * rate
                unit_price = unit_price * rate

            position = Position(
                symbol=symbol,
                quantity=quantity,
                position_datetime=position_datetime,
                total_value=market_value,
                book_value=book_value,
                unit_price=unit_price
            )
            positions.append(position)

    return positions